"""
Extraction script: Table 1 OHHLEP Tools Inventory xlsx -> content/oh-tools.json
One-shot, re-runnable, idempotent. Committed for provenance.

Reads rows 3-52 (the 50 One Health tools) from the "Table 1 OHHLEP Tools
Inventory" sheet, excluding the title (row 1), header (row 2), and the
footnote/citation rows (54-64).
"""

import hashlib
import json
import re
from pathlib import Path

import openpyxl

XLSX_PATH = (
    Path(__file__).parent
    / "Table 1_One Health Tools Inventory_13Feb2024_ohjpa_Annex2_v4-1.xlsx"
)
OUTPUT_PATH = Path(__file__).parent.parent.parent / "content" / "oh-tools.json"

# Pull the first http(s) URL or a bare www. host out of a (possibly prose) cell.
URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)
WWW_RE = re.compile(r"\bwww\.[^\s,;]+", re.IGNORECASE)


def clean(v) -> str:
    """Collapse embedded newlines / runs of whitespace to single spaces."""
    return re.sub(r"\s+", " ", str(v).strip()) if v is not None else ""


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return text


def fix_url(url_val: str) -> str:
    """Normalize a messy C13 cell to a single https URL or empty string.

    - Returns "" for empty input.
    - Returns the first https?://... token if a URL is buried in prose.
    - Prefixes a bare www. / scheme-less host with https://.
    - Returns "" if no URL-like token is found.
    """
    url_val = (url_val or "").strip()
    if not url_val:
        return ""

    # Buried/explicit http(s) URL anywhere in the cell.
    m = URL_RE.search(url_val)
    if m:
        return m.group(0).rstrip(".,;")

    # Bare www. host.
    m = WWW_RE.search(url_val)
    if m:
        return "https://" + m.group(0).rstrip(".,;")

    # Scheme-less host fragment with a dot and a path (e.g. "Github.com/EIDSS").
    # Guard against false positives (version strings, decimal numbers, abbreviations).
    m = re.search(r"\b[\w.-]+\.[a-z]{2,}(?:/\S*)?", url_val, re.IGNORECASE)
    if m:
        candidate = "https://" + m.group(0).rstrip(".,;")
        # Only accept if it looks like a real host (has a path or a proper TLD suffix)
        if re.search(r"\.[a-z]{2,}/", candidate) or re.search(r"\.[a-z]{2,}$", candidate):
            import sys
            print(
                f"  [fix_url] fallback regex matched '{m.group(0)}' — verify manually",
                file=sys.stderr,
            )
            return candidate

    return ""


# Canonicalize source casing/spelling variants to the verified token unions
# in app/lib/types.ts (OHScope / OHAudienceType). Keys are lowercased.
TOKEN_CANON = {
    # OHScope variants
    "prioritising": "Prioritisation",
    "prioritisation": "Prioritisation",
    "implementation": "Implementation",
    "assessment": "Assessment",
    "monitoring": "Monitoring",
    "action plans": "Action Plans",
    # OHAudienceType variants
    "terrestrial and aquatic animal health and welfare": "Animal health",
}


def split_clean(v) -> list[str]:
    """Split a comma-combined cell, clean each token, drop empties.

    Normalizes the wrapped "International/\nRegional" value to
    "International/Regional" and canonicalizes casing/spelling variants
    (e.g. "Prioritising"->"Prioritisation", "implementation"->"Implementation")
    to the verified token unions so the JSON validates against ToolItem.
    """
    raw = clean(v)
    if not raw:
        return []
    out: list[str] = []
    for part in raw.split(","):
        token = part.strip()
        token = re.sub(r"International/\s*Regional", "International/Regional", token)
        if not token:
            continue
        token = TOKEN_CANON.get(token.lower(), token)
        out.append(token)
    return out


def build_items() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Table 1 OHHLEP Tools Inventory"]

    seen_ids: set[str] = set()
    items: list[dict] = []

    for r in range(3, 53):  # rows 3-52 ONLY (exclude footnotes 54-64)
        name = clean(ws.cell(r, 1).value)
        year_val = ws.cell(r, 2).value
        # Defense-in-depth against footnotes: require a name AND a year.
        if not name or year_val is None:
            continue

        organization = clean(ws.cell(r, 3).value)
        organization_levels = split_clean(ws.cell(r, 4).value)
        scopes = split_clean(ws.cell(r, 5).value)
        audience_levels = [t.title() for t in split_clean(ws.cell(r, 8).value)]
        audience_types = split_clean(ws.cell(r, 9).value)
        description = clean(ws.cell(r, 11).value)
        url = fix_url(clean(ws.cell(r, 13).value))

        # Build slug id; dedup with a content-hash suffix so IDs are stable
        # across re-runs regardless of row order.
        base_slug = slugify(name)[:60]
        slug = base_slug
        if slug in seen_ids:
            h = hashlib.md5(name.encode()).hexdigest()[:6]
            slug = f"{base_slug}-{h}"
        seen_ids.add(slug)

        items.append(
            {
                "id": slug,
                "name": name,
                "year": int(year_val),
                "organization": organization,
                "organizationLevels": organization_levels,
                "scopes": scopes,
                "audienceLevels": audience_levels,
                "audienceTypes": audience_types,
                "description": description,
                "url": url,
            }
        )

    return items


if __name__ == "__main__":
    items = build_items()
    print(len(items))

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
        f.write("\n")
